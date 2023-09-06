# -*- coding: utf-8 -*-
# time: 2023/7/13 18:12
# file: main.py
# author: xiong
import sys
import numpy as np
import openpyxl
from handle import detect_circle

sys.path.append("camera")
from camera.driver.MvCameraControl_class import *



    while num <= 5:
        # 调用相机拍照并保存ITEST.jpg
        deviceList = MV_CC_DEVICE_INFO_LIST()
        tlayerType = MV_GIGE_DEVICE | MV_USB_DEVICE

        # ch:枚举设备 | en:Enum device
        ret = MvCamera.MV_CC_EnumDevices(tlayerType, deviceList)
        if ret != 0:
            print("enum devices fail! ret[0x%x]" % ret)
            sys.exit()

        if deviceList.nDeviceNum == 0:
            print("find no device!")
            sys.exit()

        print("find %d devices!" % deviceList.nDeviceNum)

        for i in range(0, deviceList.nDeviceNum):
            mvcc_dev_info = cast(deviceList.pDeviceInfo[i], POINTER(MV_CC_DEVICE_INFO)).contents
            if mvcc_dev_info.nTLayerType == MV_GIGE_DEVICE:
                print("\ngige device: [%d]" % i)
                strModeName = ""
                for per in mvcc_dev_info.SpecialInfo.stGigEInfo.chModelName:
                    strModeName = strModeName + chr(per)
                print("device model name: %s" % strModeName)

                nip1 = ((mvcc_dev_info.SpecialInfo.stGigEInfo.nCurrentIp & 0xff000000) >> 24)
                nip2 = ((mvcc_dev_info.SpecialInfo.stGigEInfo.nCurrentIp & 0x00ff0000) >> 16)
                nip3 = ((mvcc_dev_info.SpecialInfo.stGigEInfo.nCurrentIp & 0x0000ff00) >> 8)
                nip4 = (mvcc_dev_info.SpecialInfo.stGigEInfo.nCurrentIp & 0x000000ff)
                print("current ip: %d.%d.%d.%d\n" % (nip1, nip2, nip3, nip4))
            elif mvcc_dev_info.nTLayerType == MV_USB_DEVICE:
                print("\nu3v device: [%d]" % i)
                strModeName = ""
                for per in mvcc_dev_info.SpecialInfo.stUsb3VInfo.chModelName:
                    if per == 0:
                        break
                    strModeName = strModeName + chr(per)
                print("device model name: %s" % strModeName)

                strSerialNumber = ""
                for per in mvcc_dev_info.SpecialInfo.stUsb3VInfo.chSerialNumber:
                    if per == 0:
                        break
                    strSerialNumber = strSerialNumber + chr(per)
                print("user serial number: %s" % strSerialNumber)

        nConnectionNum = 0

        if int(nConnectionNum) >= deviceList.nDeviceNum:
            print("intput error!")
            sys.exit()

        # ch:创建相机实例 | en:Creat Camera Object
        cam = MvCamera()

        # ch:选择设备并创建句柄 | en:Select device and create handle
        stDeviceList = cast(deviceList.pDeviceInfo[int(nConnectionNum)], POINTER(MV_CC_DEVICE_INFO)).contents

        ret = cam.MV_CC_CreateHandle(stDeviceList)
        if ret != 0:
            print("create handle fail! ret[0x%x]" % ret)
            sys.exit()

        # ch:打开设备 | en:Open device
        ret = cam.MV_CC_OpenDevice(MV_ACCESS_Exclusive, 0)
        if ret != 0:
            print("open device fail! ret[0x%x]" % ret)
            sys.exit()

        # ch:探测网络最佳包大小(只对GigE相机有效) | en:Detection network optimal package size(It only works for the GigE camera)
        if stDeviceList.nTLayerType == MV_GIGE_DEVICE:
            nPacketSize = cam.MV_CC_GetOptimalPacketSize()
            if int(nPacketSize) > 0:
                ret = cam.MV_CC_SetIntValue("GevSCPSPacketSize", nPacketSize)
                if ret != 0:
                    print("Warning: Set Packet Size fail! ret[0x%x]" % ret)
            else:
                print("Warning: Get Packet Size fail! ret[0x%x]" % nPacketSize)

        # ch:设置触发模式为off | en:Set trigger mode as off
        ret = cam.MV_CC_SetEnumValue("TriggerMode", MV_TRIGGER_MODE_OFF)
        if ret != 0:
            print("set trigger mode fail! ret[0x%x]" % ret)
            sys.exit()

        # ch:设置像素格式 | en:Set pixel type (New)
        ret = cam.MV_CC_SetEnumValue("PixelFormat", PixelType_Gvsp_Mono12)
        if ret != 0:
            print("set pixel type fail! ret[0x%x]" % ret)
            sys.exit()

        # ch:获取数据包大小 | en:Get payload size
        stParam = MVCC_INTVALUE()
        memset(byref(stParam), 0, sizeof(MVCC_INTVALUE))

        ret = cam.MV_CC_GetIntValue("PayloadSize", stParam)
        if ret != 0:
            print("get payload size fail! ret[0x%x]" % ret)
            sys.exit()

        nPayloadSize = stParam.nCurValue

        # ch:开始取流 | en:Start grab image
        ret = cam.MV_CC_StartGrabbing()
        if ret != 0:
            print("start grabbing fail! ret[0x%x]" % ret)
            sys.exit()

        stDeviceList = MV_FRAME_OUT_INFO_EX()
        memset(byref(stDeviceList), 0, sizeof(stDeviceList))
        # data_buf即为所求的RAW文件
        data_buf = (c_ubyte * nPayloadSize)()

        ret = cam.MV_CC_GetOneFrameTimeout(byref(data_buf), nPayloadSize, stDeviceList, 1000)
        if ret == 0:
            print("\nget one frame: Width[%d], Height[%d], nFrameNum[%d]\n" % (
                stDeviceList.nWidth, stDeviceList.nHeight, stDeviceList.nFrameNum))

        # ch:亮度检测 | en:Detect luminance
        img = np.frombuffer(data_buf, dtype=np.int16)
        img = img.reshape(stDeviceList.nHeight, stDeviceList.nWidth)
        luminance_avg = detect_circle.detect(img)  # 亮度计算
        values.append(round(luminance_avg, 2))

        # ch:停止取流 | en:Stop grab image
        # ch:将raw文件转换位uint16位数组 | en:Convert raw into uint16 array

        ret = cam.MV_CC_StopGrabbing()
        if ret != 0:
            print("stop grabbing fail! ret[0x%x]" % ret)
            del data_buf
            sys.exit()

        # ch:关闭设备 | Close device
        ret = cam.MV_CC_CloseDevice()
        if ret != 0:
            print("close deivce fail! ret[0x%x]" % ret)
            del data_buf
            sys.exit()

        # ch:销毁句柄 | Destroy handle
        ret = cam.MV_CC_DestroyHandle()
        if ret != 0:
            print("destroy handle fail! ret[0x%x]" % ret)
            del data_buf
            sys.exit()

        del data_buf
        num += 1

    # 将亮度写入excel表中
    workbook = openpyxl.load_workbook('output/luminance_avg.xlsx')
    worksheet = workbook.sheetnames  # 取第一张表
    worktable = workbook[worksheet[0]]  # 获取第一列
    rows = worktable.max_row  # 获得行数
    for value in values:
        worktable.cell(rows + 1, 1).value = value
        rows = rows + 1
    workbook.save('../output/luminance_avg.xls')  # 将新数据追加进excel表中
    workbook.close()
