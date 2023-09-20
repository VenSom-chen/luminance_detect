import numpy as np
import openpyxl
from PyQt5.QtCore import QObject, pyqtSignal, Qt
from PyQt5.QtGui import QIntValidator

from app.controller.common.mbox import MessageBox
from app.controller.common.rawpix import RawPix
from app.controller.thread.camera_manager import CameraManager
from app.gui.main_win import main_window
from handle import detect_circle


class MainWinController(QObject):
    video_acted = pyqtSignal(object)
    show_value = pyqtSignal()
    auto_stopped = pyqtSignal(str,str)

    def __init__(self):
        super(MainWinController, self).__init__()
        # ui声明
        self.mw = main_window()  # 主界面
        self.mbox = MessageBox(self.mw)  # 消息弹窗
        # 变量声明
        self.cam = CameraManager.instance()  # 相机实例
        self.current_device = None  # 当前设备
        self.is_handling = False  # 相机正在打开时控制提示
        # 数据处理
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.value_in_col = []
        self.auto_started = None
        self.col = 1  # 列的初始值
        self.row = 1  # 行的初始值
        self.file_path = '../output/luminance_avg.xlsx'
        self.time = 100  # 自动检测次数
        self.current_luminance = None
        self.init_ui()
        self.init_camera()
        self.video_acted.connect(self.video_show, Qt.QueuedConnection)
        self.auto_stopped.connect(self.refresh_ui)

    # 初始化ui
    def init_ui(self):
        # mw组件初始化
        self.mw.time_set.setValidator(QIntValidator(0, 1000))
        self.mw.auto_star.setDisabled(True)
        self.mw.photo_show.rawpix = RawPix(self.mw.photo_show)
        self.mw.open_cam.clicked.connect(self.open_camera)
        self.mw.close_cam.clicked.connect(self.close_camera)
        self.mw.refresh_list.clicked.connect(self.refresh_combox)
        self.mw.auto_star.clicked.connect(self.auto_detect)
        self.mw.time_set.setText('100')

    # 初始化相机
    def init_camera(self):
        self.cam.camera_opened_connect(self.camera_opened)
        self.cam.camera_closed_connect(self.camera_closed)
        self.cam.exception_handle_connect(self.exception_handle)
        self.cam.devices_changed_connect(self.combox_refreshed)


    def video_show(self, image):
        self.mw.photo_show.rawpix.set_pixmap_with_cvimg(image)

    # 打开相机
    def open_camera(self):
        if self.is_handling:
            self.mbox.warn("警告", "请勿操作过快")
            return
        self.is_handling = True
        device_name = self.mw.cam_combox.currentText()
        if len(device_name) == 0:
            self.mbox.info("提示", "未发现相机")
            self.is_handling = False
            return
        exposure_time = self.mw.exposure_time_edit.text()
        if len(exposure_time) == 0:
            self.mbox.info("提示", "未设置曝光时间")
            self.is_handling = False
            return
        self.cam.open_camera(exposure_time, device_name, work=self.work_thread)
        self.mw.cam_combox.setDisabled(False)

    # 相机打开的槽函数
    def camera_opened(self):
        print("相机已打开")
        self.is_handling = False
        self.current_device = self.mw.cam_combox.currentText()
        self.mw.cam_combox.setDisabled(True)
        self.mw.auto_star.setDisabled(False)

    def close_camera(self):
        if self.is_handling:
            self.mbox.warn("警告","请勿操作过快")
            return
        self.is_handling = True
        self.cam.close_camera(self.current_device)
        self.mw.cam_combox.setDisabled(False)

    # 相机关闭的参函数
    def camera_closed(self):
        print("相机已关闭")
        self.current_device = None
        self.is_handling = False
        self.mw.cam_combox.setDisabled(False)
        self.mw.auto_star.setDisabled(True)
        self.mw.photo_show.clear()

    def refresh_combox(self):
        camera_list = self.cam.refresh_devices()
        self.mw.cam_combox.clear()
        self.mw.cam_combox.addItems(camera_list)

    def auto_detect(self):
        if self.is_handling:
            self.mbox.warn("警告", "请勿操作过快")
            return
        self.is_handling = True
        self.auto_started = True
        self.mw.auto_star.setDisabled(True)
        self.time = int(self.mw.time_set.text())

    # 相机工作
    def work_thread(self, device_name, data, size_info):
        image = np.frombuffer(data, dtype=np.uint16)  # 将c_ubyte_Array转化成ndarray得到（3686400，）
        image = image.reshape(size_info.nHeight, size_info.nWidth)  # 根据自己分辨率进行转化
        try:
            image, lumi_avg = detect_circle.detect(image)  # 亮度计算
        except Exception as e:
            print(e)
            self.video_acted.emit(image)
            return
        self.video_acted.emit(image)
        # 如果自动操作模式打开则记录数据
        if not self.auto_started:
            return
        if self.row <= self.time:
            self.worksheet.cell(row=self.row, column=self.col, value=lumi_avg)
            self.value_in_col.append(lumi_avg)
            self.row += 1
            return
        # 写100个数据后
        # 写入平均值
        mean = format(np.array(self.value_in_col).mean(),'.2f')
        print(mean)
        self.worksheet.cell(row=self.row, column=self.col,
                            value=mean)
        # 写入方差
        var = format(np.array(self.value_in_col).var(),'.2f')
        self.worksheet.cell(row=self.row + 1, column=self.col,
                            value=var)
        # 还原
        self.workbook.save(self.file_path)
        self.value_in_col.clear()
        self.row = 1
        self.col += 1
        self.auto_started = False
        self.auto_stopped.emit(mean,var)

    def refresh_ui(self,mean,var):
        self.mw.blur_value.setText(mean)
        self.mw.var_value.setText(var)
        self.mw.auto_star.setDisabled(False)
        self.is_handling = False
        self.mbox.info("提示","自动检测完成")

    # 相机列表跟新槽函数
    def combox_refreshed(self, camera_list):
        self.mw.cam_combox.clear()
        self.mw.cam_combox.addItems(camera_list)

    def exception_handle(self, e):
        self.is_handling = False
        self.mw.cam_combox.setDisabled(False)
        self.mbox.error('错误', str(e))
